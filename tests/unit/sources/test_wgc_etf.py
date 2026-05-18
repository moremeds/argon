"""WGC ETF monthly workbook provider."""

from __future__ import annotations

from datetime import date
from decimal import Decimal
from io import BytesIO

import httpx
from openpyxl import Workbook

from uw_scan.sources.wgc_etf import TROY_OZ_PER_TONNE, WgcEtfProvider


def _workbook_bytes() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Holdings by month"
    ws.append(
        [
            "ticker",
            "All units in tonnes unless otherwise specified",
            None,
            None,
            None,
            "gld us equity",
            "iau us equity",
            "gldm us equity",
            "phys us equity",
        ]
    )
    ws.append(["Active", None, None, None, None, "Active", "Active", "Active", "Active"])
    ws.append(["Fund Type", None, None, None, None, "ETF", "ETF", "ETF", "Closed-End Fund"])
    ws.append(["Region", None, None, None, None, "North America", "North America", "North America", "North America"])
    ws.append(["Country", None, None, None, None, "US", "US", "US", "US"])
    ws.append(
        [
            "Date",
            "Gold, US$/oz",
            "Ounces",
            "Tonnes",
            "Value (USD)",
            "SPDR Gold Shares",
            "iShares Gold Trust",
            "SPDR Gold MiniShares Trust",
            "Sprott Physical Gold Trust",
        ]
    )
    ws.append(
        [
            date(2026, 2, 28),
            Decimal("2894.725"),
            None,
            None,
            None,
            Decimal("1100.0"),
            Decimal("450.0"),
            Decimal("190.0"),
            Decimal("100.0"),
        ]
    )
    ws.append(
        [
            date(2026, 3, 31),
            Decimal("2983.25"),
            None,
            None,
            None,
            Decimal("1046.9009356"),
            Decimal("475.96149609"),
            Decimal("199.89033267"),
            Decimal("114.72237255"),
        ]
    )
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _full_monthly_workbook_bytes() -> bytes:
    wb = Workbook()
    for idx, title in enumerate(
        ["Holdings by month", "Demand by month", "Fund flows by month"]
    ):
        ws = wb.active if idx == 0 else wb.create_sheet(title)
        ws.title = title
        ws.append(
            [
                "ticker",
                "All units in tonnes unless otherwise specified",
                None,
                None,
                None,
                "gld us equity",
                "iau us equity",
            ]
        )
        ws.append(["Active", None, None, None, None, "Active", "Active"])
        ws.append(["Fund Type", None, None, None, None, "ETF", "ETF"])
        ws.append(["Region", None, None, None, None, "North America", "North America"])
        ws.append(["Country", None, None, None, None, "US", "US"])
        ws.append(
            [
                "Date",
                "Gold, US$/oz",
                "Ounces",
                "Tonnes",
                "Value (USD)",
                "SPDR Gold Shares",
                "iShares Gold Trust",
            ]
        )
    wb["Holdings by month"].append(
        [
            date(2026, 3, 31),
            Decimal("2983.25"),
            Decimal("131428333.123"),
            Decimal("4087.79194987"),
            Decimal("606500000000"),
            Decimal("1046.9009356"),
            Decimal("475.96149609"),
        ]
    )
    wb["Demand by month"].append(
        [
            date(2026, 3, 31),
            Decimal("2983.25"),
            Decimal("131428333.123"),
            Decimal("4087.79194987"),
            Decimal("606500000000"),
            Decimal("-54.13175268"),
            Decimal("-23.26570443"),
        ]
    )
    wb["Fund flows by month"].append(
        [
            date(2026, 3, 31),
            Decimal("2983.25"),
            Decimal("131428333.123"),
            Decimal("4087.79194987"),
            Decimal("606500000000"),
            Decimal("-8426.7817"),
            Decimal("-3677.5729"),
        ]
    )
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_wgc_etf_provider_parses_monthly_holdings_for_gold_funds() -> None:
    rows = WgcEtfProvider().parse_holdings(
        _workbook_bytes(), start=date(2026, 3, 1)
    )

    assert [row.ticker for row in rows] == ["GLD", "IAU", "GLDM", "PHYS"]
    assert rows[0].obs_date == date(2026, 3, 31)
    assert rows[0].holdings_oz == Decimal("1046.9009356") * TROY_OZ_PER_TONNE
    assert rows[1].holdings_oz == Decimal("475.96149609") * TROY_OZ_PER_TONNE


def test_wgc_etf_provider_discovers_xlsx_downloads() -> None:
    provider = WgcEtfProvider()
    provider._client = httpx.Client(
        transport=httpx.MockTransport(
            lambda request: httpx.Response(
                200,
                text="""
                <a class="download xlsx private"
                   href="/download/file/20717/ETF_Flows_March_2026.xlsx">Data</a>
                <a href="/download/file/20718/Gold_ETF_Commentary_March_2026.pdf">PDF</a>
                """,
                request=request,
            )
        )
    )
    try:
        downloads = provider.fetch_downloads()
    finally:
        provider.close()

    assert len(downloads) == 1
    assert downloads[0].url == (
        "https://www.gold.org/download/file/20717/ETF_Flows_March_2026.xlsx"
    )


def test_wgc_etf_provider_parses_monthly_holdings_demand_and_flows() -> None:
    rows = WgcEtfProvider().parse_monthly_rows(
        _full_monthly_workbook_bytes(),
        source_url="https://www.gold.org/download/file/20717/ETF_Flows_March_2026.xlsx",
        source_label="ETF_Flows_March_2026.xlsx",
        start=date(2026, 3, 1),
    )

    assert len(rows) == 2
    assert rows[0].ticker == "GLD"
    assert rows[0].obs_date == date(2026, 3, 31)
    assert rows[0].fund_name == "SPDR Gold Shares"
    assert rows[0].region == "North America"
    assert rows[0].country == "US"
    assert rows[0].gold_price_usd_oz == Decimal("2983.25")
    assert rows[0].aggregate_holdings_tonnes == Decimal("4087.79194987")
    assert rows[0].aggregate_value_usd == Decimal("606500000000")
    assert rows[0].holdings_tonnes == Decimal("1046.9009356")
    assert rows[0].demand_tonnes == Decimal("-54.13175268")
    assert rows[0].flow_usd_mn == Decimal("-8426.7817")
    assert rows[0].source_url.endswith("ETF_Flows_March_2026.xlsx")
