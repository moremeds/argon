"""WGC monthly CB reserves parser."""

from __future__ import annotations

from datetime import date
from decimal import Decimal
from pathlib import Path
from unittest.mock import patch

import httpx
from openpyxl import Workbook

from uw_scan.cards.cb_buckets import classify_bucket
from uw_scan.sources.wgc_cb import WgcCbProvider

SAMPLE = """Country,Month,Tonnes,Reported,Estimated
China,2026-04,2235.0,true,false
India,2026-04,876.4,true,false
Russia,2026-04,2330.5,false,true
Poland,2026-04,420.3,true,false
"""


def _fake_response() -> httpx.Response:
    return httpx.Response(
        200,
        text=SAMPLE,
        request=httpx.Request("GET", WgcCbProvider.URL),
    )


def test_wgc_parses_monthly_csv():
    with patch.object(WgcCbProvider, "_get_with_telemetry") as mock_get:
        mock_get.return_value = _fake_response()
        with WgcCbProvider() as p:
            rows = p.fetch_monthly(start=date(2026, 4, 1))
    by_country = {r.country_iso3: r for r in rows}
    assert by_country["CHN"].reserves_t == Decimal("2235.0")
    assert by_country["CHN"].obs_month == date(2026, 4, 1)
    assert by_country["RUS"].is_reported is False
    assert by_country["RUS"].is_estimated is True
    assert by_country["POL"].bucket == "reserve_diversifier"
    assert by_country["CHN"].bucket == "strategic_accumulator"


def test_wgc_parses_quarterly_workbook(tmp_path: Path):
    workbook_path = tmp_path / "Quarterly_gold_and_FX_Reserves_Q1_2026.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Gold (Tonnes)"
    ws.cell(row=2, column=3, value="Q4 2025")
    ws.cell(row=2, column=4, value="Q1 2026")
    ws.cell(row=3, column=1, value="China, P.R.: Mainland")
    ws.cell(row=3, column=2, value="China, P.R.: Mainland")
    ws.cell(row=3, column=3, value=2306.304625)
    ws.cell(row=3, column=4, value=2313.458368)
    ws.cell(row=4, column=1, value="Türkiye, Rep of5")
    ws.cell(row=4, column=2, value="Türkiye, Rep of5")
    ws.cell(row=4, column=3, value=614.302763)
    ws.cell(row=4, column=4, value=534.851285)
    wb.save(workbook_path)

    rows = WgcCbProvider.parse_quarterly_workbook(
        workbook_path, start=date(2026, 1, 1)
    )

    by_country = {r.country_iso3: r for r in rows}
    assert by_country["CHN"].obs_month == date(2026, 3, 31)
    assert by_country["CHN"].reserves_t == Decimal("2313.458368")
    assert by_country["CHN"].bucket == "strategic_accumulator"
    assert by_country["TUR"].reserves_t == Decimal("534.851285")


def test_classify_bucket_defaults():
    assert classify_bucket("CHN") == "strategic_accumulator"
    assert classify_bucket("EGY") == "tactical_defender"
    assert classify_bucket("POL") == "reserve_diversifier"
    assert classify_bucket("USA") == "reserve_diversifier"  # safe default
