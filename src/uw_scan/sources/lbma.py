"""LBMA monthly vault holdings (loco London).

Source: https://www.lbma.org.uk/prices-and-data/london-vault-data
LBMA moved from a stable .csv URL to monthly-named .xlsx files hosted at
cdn.lbma.org.uk/downloads/LBMA-London-Vault-Holdings-Data-<Month-Year>.xlsx.
We scrape the listing page each run to discover the current URL, then parse
the workbook. Gold column is reported in thousands of troy ounces — we
multiply by 1000 so vault_oz stays in oz (consistent with COMEX).
"""

from __future__ import annotations

import io
import logging
import re
from collections.abc import Callable
from dataclasses import dataclass
from datetime import UTC, date, datetime
from decimal import Decimal, InvalidOperation
from typing import Any

import httpx
from openpyxl import load_workbook

from uw_scan.storage.provider_usage import ExternalApiRequestEvent
from uw_scan.storage.repository import redact_params, status_family_for

logger = logging.getLogger(__name__)


@dataclass(frozen=True)
class LbmaVaultRow:
    obs_date: date
    vault_oz: Decimal | None


RecordHook = Callable[["LbmaProvider", ExternalApiRequestEvent], None]

_XLSX_URL_RE = re.compile(
    r"https://cdn\.lbma\.org\.uk/downloads/"
    r"LBMA-London-Vault-Holdings-Data-[A-Za-z]+-\d{4}\.xlsx",
    re.IGNORECASE,
)


class LbmaProvider:
    LISTING_URL = "https://www.lbma.org.uk/prices-and-data/london-vault-data"
    URL = LISTING_URL
    ENDPOINT_PATH = "/prices-and-data/london-vault-data"
    ENDPOINT_KEY = "lbma_vault_xlsx"
    PROVIDER = "lbma"
    BROWSER_UA = (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 13_0) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
    )
    OZ_PER_THOUSAND = Decimal(1000)

    def __init__(
        self,
        *,
        timeout_s: float = 30.0,
        record_request: RecordHook | None = None,
    ):
        self._client = httpx.Client(
            timeout=timeout_s, headers={"User-Agent": self.BROWSER_UA}
        )
        self._record_request_fn = record_request

    def close(self) -> None:
        self._client.close()

    def __enter__(self) -> "LbmaProvider":
        return self

    def __exit__(self, *_exc: object) -> None:
        self.close()

    def fetch_monthly(self, *, start: date | None = None) -> list[LbmaVaultRow]:
        listing = self._get_with_telemetry(self.LISTING_URL, {})
        listing.raise_for_status()
        xlsx_urls = sorted(set(_XLSX_URL_RE.findall(listing.text)))
        if not xlsx_urls:
            logger.warning("lbma: no vault-holdings xlsx URL found on listing page")
            return []
        # The newest file is the one we want; the listing typically shows the
        # most recent first, but sort defensively (filename includes Month-Year).
        xlsx_url = xlsx_urls[-1]
        workbook_response = self._get_with_telemetry(xlsx_url, {})
        workbook_response.raise_for_status()
        return _parse_workbook(workbook_response.content, start=start)

    def _get_with_telemetry(self, url: str, params: dict[str, Any]) -> httpx.Response:
        started_at = datetime.now(UTC)
        try:
            response = self._client.get(url, params=params)
        except httpx.HTTPError as exc:
            finished_at = datetime.now(UTC)
            self._record_request(
                self._build_event(
                    started_at,
                    finished_at,
                    params,
                    status_code=None,
                    error_message=repr(exc)[:1000],
                )
            )
            raise
        finished_at = datetime.now(UTC)
        self._record_request(
            self._build_event(
                started_at,
                finished_at,
                params,
                status_code=response.status_code,
                error_message=(
                    response.text[:1000] if response.status_code >= 400 else None
                ),
            )
        )
        return response

    def _record_request(self, event: ExternalApiRequestEvent) -> None:
        if self._record_request_fn is not None:
            self._record_request_fn(self, event)
        else:
            logger.debug("lbma telemetry %r", event)

    def _build_event(
        self,
        started_at: datetime,
        finished_at: datetime,
        params: dict[str, Any],
        *,
        status_code: int | None,
        error_message: str | None,
    ) -> ExternalApiRequestEvent:
        return ExternalApiRequestEvent(
            provider=self.PROVIDER,
            endpoint_key=self.ENDPOINT_KEY,
            method="GET",
            path=self.ENDPOINT_PATH,
            path_template=self.ENDPOINT_PATH,
            params=redact_params(params),
            status_code=status_code,
            status_family=status_family_for(
                status_code, transport_error=status_code is None
            ),
            started_at=started_at,
            finished_at=finished_at,
            latency_ms=max(0, int((finished_at - started_at).total_seconds() * 1000)),
            error_message=error_message,
        )


def _parse_workbook(content: bytes, *, start: date | None) -> list[LbmaVaultRow]:
    """Parse the LBMA monthly vault-holdings .xlsx into LbmaVaultRow.

    Sheet layout (verified 2026-05-17):
      Row 0: title "London Vault Holdings Data"
      Row 1: header ["Month End", "Gold", "Silver"]
      Row 2: units ["", "Troy Ounces ('000s)", "Troy Ounces ('000s)"]
      Row 3+: data — Month End is either "YYYY-MM" string or datetime,
              Gold/Silver are integers in thousands of troy ounces.
    """
    workbook = load_workbook(io.BytesIO(content), data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    out: list[LbmaVaultRow] = []
    for row in sheet.iter_rows(min_row=4, values_only=True):
        if not row or row[0] is None or row[1] is None:
            continue
        obs = _parse_month_end(row[0])
        if obs is None:
            continue
        if start is not None and obs < start:
            continue
        try:
            vault_oz = Decimal(str(row[1])) * LbmaProvider.OZ_PER_THOUSAND
        except (InvalidOperation, ValueError) as exc:
            logger.debug("lbma: unparseable Gold %r (%s)", row[1], repr(exc))
            continue
        out.append(LbmaVaultRow(obs_date=obs, vault_oz=vault_oz))
    return out


def _parse_month_end(raw: object) -> date | None:
    if isinstance(raw, datetime):
        return raw.date()
    if isinstance(raw, date):
        return raw
    if isinstance(raw, str):
        s = raw.strip()
        try:
            if len(s) == 7:
                return date.fromisoformat(s + "-01")
            return date.fromisoformat(s)
        except ValueError as exc:
            logger.debug("lbma: unparseable Month End %r (%s)", raw, repr(exc))
            return None
    return None
