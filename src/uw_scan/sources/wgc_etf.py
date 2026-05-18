"""World Gold Council monthly gold ETF holdings workbook."""

from __future__ import annotations

import io
import logging
from collections.abc import Iterable
from dataclasses import dataclass
from datetime import UTC, date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any
from urllib.parse import urljoin

import httpx
from bs4 import BeautifulSoup
from openpyxl import load_workbook

from uw_scan.sources.etf_holdings import EtfHoldingRow
from uw_scan.storage.provider_usage import ExternalApiRequestEvent
from uw_scan.storage.repository import redact_params, status_family_for

logger = logging.getLogger(__name__)

TROY_OZ_PER_TONNE = Decimal("32150.7466")


@dataclass(frozen=True)
class WgcEtfDownload:
    url: str
    label: str


@dataclass(frozen=True)
class WgcEtfMonthlyRow:
    ticker: str
    obs_date: date
    fund_name: str | None
    fund_type: str | None
    region: str | None
    country: str | None
    gold_price_usd_oz: Decimal | None
    aggregate_ounces: Decimal | None
    aggregate_holdings_tonnes: Decimal | None
    aggregate_value_usd: Decimal | None
    holdings_tonnes: Decimal | None
    demand_tonnes: Decimal | None
    flow_usd_mn: Decimal | None
    source_url: str
    source_label: str | None


class WgcEtfProvider:
    """Fetch and parse WGC's monthly ETF flows workbook.

    The listing page is public, but the XLSX downloads are gated by Goldhub
    login. Pass a cookie header only from local environment/config; never bake
    it into source.
    """

    ETF_FLOWS_PAGE = "https://www.gold.org/goldhub/research/etf-flows"
    PROVIDER = "wgc_etf"
    ENDPOINT_KEY = "wgc_etf_flows_xlsx"
    DEFAULT_TIMEOUT_S = 60.0
    BROWSER_UA = (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 13_0) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
    )
    DEFAULT_TICKERS = {
        "gld us equity": "GLD",
        "iau us equity": "IAU",
        "gldm us equity": "GLDM",
        "phys us equity": "PHYS",
    }

    def __init__(
        self,
        *,
        cookie_header: str | None = None,
        timeout_s: float | None = None,
        record_request: Any | None = None,
    ) -> None:
        headers = {"User-Agent": self.BROWSER_UA}
        if cookie_header:
            headers["Cookie"] = cookie_header
        self._client = httpx.Client(
            timeout=timeout_s if timeout_s is not None else self.DEFAULT_TIMEOUT_S,
            headers=headers,
            follow_redirects=True,
        )
        self._record_request_fn = record_request

    def close(self) -> None:
        self._client.close()

    def __enter__(self) -> "WgcEtfProvider":
        return self

    def __exit__(self, *_exc: object) -> None:
        self.close()

    def fetch_downloads(self, *, max_pages: int = 9) -> list[WgcEtfDownload]:
        downloads: list[WgcEtfDownload] = []
        seen: set[str] = set()
        for page in range(max_pages):
            url = self.ETF_FLOWS_PAGE if page == 0 else f"{self.ETF_FLOWS_PAGE}?page={page}"
            page_downloads = self._fetch_downloads_page(url)
            if not page_downloads:
                break
            for item in page_downloads:
                if item.url in seen:
                    continue
                seen.add(item.url)
                downloads.append(item)
        return downloads

    def _fetch_downloads_page(self, url: str) -> list[WgcEtfDownload]:
        response = self._get(url, endpoint_key="wgc_etf_flows_page")
        response.raise_for_status()
        soup = BeautifulSoup(response.text, "html.parser")
        downloads: list[WgcEtfDownload] = []
        for anchor in soup.select('a[href*="/download/file/"][href$=".xlsx"]'):
            href = anchor.get("href")
            if not href or "ETF" not in href.upper():
                continue
            downloads.append(
                WgcEtfDownload(
                    url=urljoin(self.ETF_FLOWS_PAGE, href),
                    label=anchor.get_text(" ", strip=True) or Path(href).name,
                )
            )
        return downloads

    def fetch_monthly_rows(
        self,
        *,
        start: date | None = None,
        max_pages: int = 9,
    ) -> list[WgcEtfMonthlyRow]:
        out: list[WgcEtfMonthlyRow] = []
        for download in self.fetch_downloads(max_pages=max_pages):
            content = self.fetch_workbook(download.url)
            out.extend(
                self.parse_monthly_rows(
                    content,
                    source_url=download.url,
                    source_label=download.label,
                    start=start,
                )
            )
        return out

    def fetch_latest_holdings(
        self,
        *,
        start: date | None = None,
        tickers: Iterable[str] | None = None,
    ) -> list[EtfHoldingRow]:
        downloads = self.fetch_downloads()
        if not downloads:
            return []
        content = self.fetch_workbook(downloads[0].url)
        return self.parse_holdings(content, start=start, tickers=tickers)

    def fetch_workbook(self, url: str) -> bytes:
        response = self._get(url, endpoint_key=self.ENDPOINT_KEY)
        response.raise_for_status()
        content_type = response.headers.get("content-type", "").lower()
        if not (
            response.content.startswith(b"PK\x03\x04")
            or "spreadsheetml.sheet" in content_type
        ):
            raise ValueError(f"WGC ETF response is not an XLSX workbook: {content_type}")
        return response.content

    def parse_holdings(
        self,
        content: bytes,
        *,
        start: date | None = None,
        tickers: Iterable[str] | None = None,
    ) -> list[EtfHoldingRow]:
        wanted = {t.upper() for t in tickers} if tickers is not None else None
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        sheet = _sheet_by_names(workbook, ["Holdings by month", "All holdings by month"])
        if sheet is None:
            logger.warning("WGC ETF workbook missing Holdings by month sheet")
            return []
        rows = list(sheet.iter_rows(values_only=True))
        if len(rows) < 7:
            return []

        ticker_row = rows[0]
        fund_row = rows[5]
        columns: list[tuple[int, str]] = []
        for idx, raw_ticker in enumerate(ticker_row):
            ticker = self.DEFAULT_TICKERS.get(str(raw_ticker or "").strip().lower())
            if ticker is None:
                continue
            if wanted is not None and ticker not in wanted:
                continue
            columns.append((idx, ticker))

        out: list[EtfHoldingRow] = []
        for row in rows[6:]:
            obs_date = _parse_excel_date(row[0] if row else None)
            if obs_date is None or (start and obs_date < start):
                continue
            for idx, ticker in columns:
                tonnes = _dec(_cell(row, idx))
                if tonnes is None:
                    continue
                out.append(
                    EtfHoldingRow(
                        ticker=ticker,
                        obs_date=obs_date,
                        holdings_oz=tonnes * TROY_OZ_PER_TONNE,
                        shares_out=None,
                        nav_per_share=None,
                        premium_pct=None,
                    )
                )
        logger.debug(
            "parsed %s WGC ETF holdings rows from columns=%s",
            len(out),
            [fund_row[idx] for idx, _ticker in columns],
        )
        return out

    def parse_monthly_rows(
        self,
        content: bytes,
        *,
        source_url: str,
        source_label: str | None = None,
        start: date | None = None,
    ) -> list[WgcEtfMonthlyRow]:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        holdings_sheet = _sheet_by_names(
            workbook, ["Holdings by month", "All holdings by month"]
        )
        if holdings_sheet is None:
            fund_flow_sheet = _sheet_by_names(
                workbook, ["All fund flows by fund", "All fund flows"]
            )
            if fund_flow_sheet is None:
                logger.warning("WGC ETF workbook missing monthly holdings sheet")
                return []
            return _parse_current_fund_flow_rows(
                list(fund_flow_sheet.iter_rows(values_only=True)),
                source_url=source_url,
                source_label=source_label,
                start=start,
            )
        demand_sheet = _sheet_by_names(
            workbook, ["Demand by month", "Delta tonnes by month"]
        )
        flow_sheet = _sheet_by_names(
            workbook, ["Fund flows by month", "All flows US$ by month"]
        )

        holdings_rows = list(holdings_sheet.iter_rows(values_only=True))
        if len(holdings_rows) < 7:
            return []

        columns = _monthly_columns(holdings_rows)
        demand_by_date = (
            _monthly_values_by_date(list(demand_sheet.iter_rows(values_only=True)))
            if demand_sheet is not None
            else {}
        )
        flows_by_date = (
            _monthly_values_by_date(list(flow_sheet.iter_rows(values_only=True)))
            if flow_sheet is not None
            else {}
        )

        out: list[WgcEtfMonthlyRow] = []
        for row in holdings_rows[6:]:
            obs_date = _parse_excel_date(_cell(row, 0))
            if obs_date is None or (start and obs_date < start):
                continue
            demand_for_date = demand_by_date.get(obs_date, {})
            flows_for_date = flows_by_date.get(obs_date, {})
            for idx, meta in columns.items():
                holdings_tonnes = _dec(_cell(row, idx))
                demand_tonnes = demand_for_date.get(idx)
                flow_usd_mn = flows_for_date.get(idx)
                if (
                    holdings_tonnes is None
                    and demand_tonnes is None
                    and flow_usd_mn is None
                ):
                    continue
                out.append(
                    WgcEtfMonthlyRow(
                        ticker=meta["ticker"],
                        obs_date=obs_date,
                        fund_name=meta["fund_name"],
                        fund_type=meta["fund_type"],
                        region=meta["region"],
                        country=meta["country"],
                        gold_price_usd_oz=_dec(_cell(row, 1)),
                        aggregate_ounces=_dec(_cell(row, 2)),
                        aggregate_holdings_tonnes=_dec(_cell(row, 3)),
                        aggregate_value_usd=_dec(_cell(row, 4)),
                        holdings_tonnes=holdings_tonnes,
                        demand_tonnes=demand_tonnes,
                        flow_usd_mn=flow_usd_mn,
                        source_url=source_url,
                        source_label=source_label,
                    )
                )
        return out

    def _get(self, url: str, *, endpoint_key: str) -> httpx.Response:
        started_at = datetime.now(UTC)
        try:
            response = self._client.get(url)
        except httpx.HTTPError as exc:
            finished_at = datetime.now(UTC)
            self._record_request(
                self._build_event(
                    url,
                    endpoint_key,
                    started_at,
                    finished_at,
                    status_code=None,
                    error_message=repr(exc)[:1000],
                )
            )
            raise
        finished_at = datetime.now(UTC)
        self._record_request(
            self._build_event(
                url,
                endpoint_key,
                started_at,
                finished_at,
                status_code=response.status_code,
                error_message=(
                    response.text[:1000] if response.status_code >= 400 else None
                ),
            )
        )
        return response

    def _record_request(self, event: ExternalApiRequestEvent) -> None:
        if self._record_request_fn is not None:
            self._record_request_fn(self, event)
        else:
            logger.debug("wgc_etf telemetry %r", event)

    def _build_event(
        self,
        url: str,
        endpoint_key: str,
        started_at: datetime,
        finished_at: datetime,
        *,
        status_code: int | None,
        error_message: str | None,
    ) -> ExternalApiRequestEvent:
        return ExternalApiRequestEvent(
            provider=self.PROVIDER,
            endpoint_key=endpoint_key,
            method="GET",
            path=url,
            path_template=url,
            params=redact_params({}),
            status_code=status_code,
            status_family=status_family_for(
                status_code, transport_error=status_code is None
            ),
            started_at=started_at,
            finished_at=finished_at,
            latency_ms=max(0, int((finished_at - started_at).total_seconds() * 1000)),
            error_message=error_message,
        )


def _cell(row: tuple[Any, ...], idx: int) -> Any:
    if idx >= len(row):
        return None
    return row[idx]


def _sheet_by_names(workbook: Any, names: list[str]) -> Any | None:
    for name in names:
        if name in workbook.sheetnames:
            return workbook[name]
    return None


def _monthly_columns(rows: list[tuple[Any, ...]]) -> dict[int, dict[str, str | None]]:
    tickers = rows[0]
    fund_types = rows[2]
    regions = rows[3]
    countries = rows[4]
    fund_names = rows[5]
    columns: dict[int, dict[str, str | None]] = {}
    for idx in range(5, len(tickers)):
        raw_ticker = str(tickers[idx] or "").strip()
        if not raw_ticker:
            continue
        columns[idx] = {
            "ticker": _display_ticker(raw_ticker),
            "fund_name": _str_or_none(_cell(fund_names, idx)),
            "fund_type": _str_or_none(_cell(fund_types, idx)),
            "region": _str_or_none(_cell(regions, idx)),
            "country": _str_or_none(_cell(countries, idx)),
        }
    return columns


def _monthly_values_by_date(rows: list[tuple[Any, ...]]) -> dict[date, dict[int, Decimal]]:
    out: dict[date, dict[int, Decimal]] = {}
    for row in rows[6:]:
        obs_date = _parse_excel_date(_cell(row, 0))
        if obs_date is None:
            continue
        values: dict[int, Decimal] = {}
        for idx in range(5, len(row)):
            value = _dec(_cell(row, idx))
            if value is not None:
                values[idx] = value
        out[obs_date] = values
    return out


def _parse_current_fund_flow_rows(
    rows: list[tuple[Any, ...]],
    *,
    source_url: str,
    source_label: str | None,
    start: date | None,
) -> list[WgcEtfMonthlyRow]:
    if len(rows) < 4:
        return []
    obs_date = _parse_as_of_date(_cell(rows[1], 1))
    if obs_date is None or (start and obs_date < start):
        return []
    out: list[WgcEtfMonthlyRow] = []
    current_region: str | None = None
    for row in rows[3:]:
        region = _str_or_none(_cell(row, 1))
        if region is not None:
            current_region = region
        raw_ticker = _str_or_none(_cell(row, 3))
        if raw_ticker is None:
            continue
        holdings_tonnes = _dec(_cell(row, 5))
        demand_tonnes = _dec(_cell(row, 8))
        flow_usd_mn = _dec(_cell(row, 9))
        if holdings_tonnes is None and demand_tonnes is None and flow_usd_mn is None:
            continue
        out.append(
            WgcEtfMonthlyRow(
                ticker=_display_ticker(raw_ticker),
                obs_date=obs_date,
                fund_name=_str_or_none(_cell(row, 2)),
                fund_type=None,
                region=current_region,
                country=_str_or_none(_cell(row, 4)),
                gold_price_usd_oz=None,
                aggregate_ounces=None,
                aggregate_holdings_tonnes=None,
                aggregate_value_usd=None,
                holdings_tonnes=holdings_tonnes,
                demand_tonnes=demand_tonnes,
                flow_usd_mn=flow_usd_mn,
                source_url=source_url,
                source_label=source_label,
            )
        )
    return out


def _parse_as_of_date(raw: Any) -> date | None:
    text = str(raw or "")
    _, _, value = text.partition("As Of Date")
    return _parse_excel_date(value.strip() or text.strip())


def _display_ticker(raw: str) -> str:
    mapped = WgcEtfProvider.DEFAULT_TICKERS.get(raw.strip().lower())
    if mapped is not None:
        return mapped
    return raw.strip().upper()


def _str_or_none(raw: Any) -> str | None:
    if raw is None:
        return None
    text = str(raw).strip()
    return text or None


def _parse_excel_date(raw: Any) -> date | None:
    if raw is None:
        return None
    if isinstance(raw, datetime):
        return raw.date()
    if isinstance(raw, date):
        return raw
    text = str(raw).strip()
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError as exc:
            logger.debug("wgc etf date parse fmt=%s skipped: %s", fmt, repr(exc))
            continue
    return None


def _dec(raw: Any) -> Decimal | None:
    if raw is None or raw == "":
        return None
    try:
        return Decimal(str(raw).replace(",", "").strip())
    except (InvalidOperation, ValueError) as exc:
        logger.debug("wgc etf decimal parse skipped: %s", repr(exc))
        return None
