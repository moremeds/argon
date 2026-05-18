"""World Gold Council central-bank gold reserves.

WGC's old anonymous monthly CSV retired in May 2026. The current Goldhub page
publishes authenticated XLSX downloads sourced from IMF IFS plus WGC
adjustments. This provider keeps the old CSV parser for historical fixtures
and adds the authenticated/local quarterly workbook path used by Goldhub.
"""

from __future__ import annotations

import csv
import io
import logging
import re
from collections.abc import Callable
from dataclasses import dataclass
from datetime import UTC, date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

import httpx
from openpyxl import load_workbook

from uw_scan.cards.cb_buckets import classify_bucket
from uw_scan.storage.provider_usage import ExternalApiRequestEvent
from uw_scan.storage.repository import redact_params, status_family_for

logger = logging.getLogger(__name__)

COUNTRY_ISO3 = {
    "argentina": "ARG",
    "australia": "AUS",
    "austria": "AUT",
    "azerbaijan": "AZE",
    "azerbaijan, rep. of": "AZE",
    "brazil": "BRA",
    "china": "CHN",
    "china, p.r.: mainland": "CHN",
    "czech rep.": "CZE",
    "india": "IND",
    "russia": "RUS",
    "russian federation": "RUS",
    "turkey": "TUR",
    "türkiye": "TUR",
    "türkiye, rep of5": "TUR",
    "poland": "POL",
    "poland, rep. of": "POL",
    "czech republic": "CZE",
    "czechia": "CZE",
    "singapore": "SGP",
    "hungary": "HUN",
    "qatar": "QAT",
    "philippines": "PHL",
    "thailand": "THA",
    "mexico": "MEX",
    "germany": "DEU",
    "france": "FRA",
    "italy": "ITA",
    "japan": "JPN",
    "united kingdom": "GBR",
    "uk": "GBR",
    "united states": "USA",
    "us": "USA",
    "switzerland": "CHE",
    "netherlands": "NLD",
    "netherlands, the": "NLD",
    "egypt": "EGY",
    "egypt, arab rep. of": "EGY",
    "kazakhstan": "KAZ",
    "kazakhstan, rep. of": "KAZ",
}


@dataclass(frozen=True)
class CbReserveRow:
    country_iso3: str
    obs_month: date
    reserves_t: Decimal | None
    bucket: str
    is_reported: bool
    is_estimated: bool


RecordHook = Callable[["WgcCbProvider", ExternalApiRequestEvent], None]


class WgcCbProvider:
    URL = "https://www.gold.org/goldhub/data/monthly-central-bank-statistics.csv"
    RESERVES_PAGE_URL = "https://www.gold.org/goldhub/data/gold-reserves-by-country"
    ENDPOINT_PATH = "/goldhub/data/monthly-central-bank-statistics.csv"
    ENDPOINT_KEY = "wgc_cb_monthly_csv"
    QUARTERLY_ENDPOINT_KEY = "wgc_cb_quarterly_xlsx"
    PROVIDER = "wgc_cb"

    def __init__(
        self,
        *,
        timeout_s: float = 30.0,
        cookie_header: str | None = None,
        workbook_path: str | Path | None = None,
        record_request: RecordHook | None = None,
    ):
        headers = {"Cookie": cookie_header} if cookie_header else None
        self._client = httpx.Client(timeout=timeout_s, headers=headers)
        self._workbook_path = Path(workbook_path) if workbook_path else None
        self._record_request_fn = record_request

    def close(self) -> None:
        self._client.close()

    def __enter__(self) -> "WgcCbProvider":
        return self

    def __exit__(self, *_exc: object) -> None:
        self.close()

    def fetch_monthly(self, *, start: date | None = None) -> list[CbReserveRow]:
        if self._workbook_path is not None:
            return self.parse_quarterly_workbook(self._workbook_path, start=start)
        if "Cookie" in self._client.headers:
            workbook_bytes, source_url = self.fetch_quarterly_workbook()
            return self.parse_quarterly_workbook(
                io.BytesIO(workbook_bytes), start=start, source_url=source_url
            )
        response = self._get_with_telemetry(self.URL, {})
        response.raise_for_status()
        return self.parse_monthly_csv(response.text, start=start)

    def fetch_quarterly_workbook(self) -> tuple[bytes, str]:
        page = self._get_with_telemetry(
            self.RESERVES_PAGE_URL,
            {},
            endpoint_key="wgc_cb_reserves_page",
            endpoint_path="/goldhub/data/gold-reserves-by-country",
        )
        page.raise_for_status()
        match = re.search(
            r'href="(?P<href>[^"]*Quarterly_gold_and_FX_Reserves[^"]*\.xlsx)"',
            page.text,
        )
        if match is None:
            raise RuntimeError("WGC CB quarterly workbook link not found")
        href = match.group("href")
        url = href if href.startswith("http") else f"https://www.gold.org{href}"
        response = self._get_with_telemetry(
            url,
            {},
            endpoint_key=self.QUARTERLY_ENDPOINT_KEY,
            endpoint_path="/download/file/:id/Quarterly_gold_and_FX_Reserves.xlsx",
        )
        response.raise_for_status()
        return response.content, url

    @classmethod
    def parse_monthly_csv(
        cls, text: str, *, start: date | None = None
    ) -> list[CbReserveRow]:
        out: list[CbReserveRow] = []
        reader = csv.DictReader(io.StringIO(text))
        for row in reader:
            country = (row.get("Country") or "").strip().lower()
            iso3 = COUNTRY_ISO3.get(country)
            if iso3 is None:
                logger.debug("wgc_cb: unknown country %r, skipping", country)
                continue
            month_raw = (row.get("Month") or "").strip()
            tonnes_raw = (row.get("Tonnes") or "").strip()
            try:
                if len(month_raw) == 7:
                    obs_month = date.fromisoformat(month_raw + "-01")
                else:
                    obs_month = date.fromisoformat(month_raw)
                reserves_t = (
                    Decimal(tonnes_raw.replace(",", "")) if tonnes_raw else None
                )
            except (ValueError, InvalidOperation) as exc:
                logger.debug("wgc cb row parse skipped: %s", repr(exc))
                continue
            if start and obs_month < start:
                continue
            is_reported = (row.get("Reported") or "").strip().lower() in (
                "true",
                "1",
                "yes",
            )
            is_estimated = (row.get("Estimated") or "").strip().lower() in (
                "true",
                "1",
                "yes",
            )
            out.append(
                CbReserveRow(
                    country_iso3=iso3,
                    obs_month=obs_month,
                    reserves_t=reserves_t,
                    bucket=classify_bucket(iso3),
                    is_reported=is_reported,
                    is_estimated=is_estimated,
                )
            )
        return out

    @classmethod
    def parse_quarterly_workbook(
        cls,
        workbook: str | Path | io.BytesIO,
        *,
        start: date | None = None,
        source_url: str | None = None,
    ) -> list[CbReserveRow]:
        del source_url
        wb = load_workbook(workbook, data_only=True, read_only=True)
        if "Gold (Tonnes)" not in wb.sheetnames:
            raise ValueError("WGC CB workbook missing Gold (Tonnes) sheet")
        sheet = wb["Gold (Tonnes)"]
        quarter_cols: list[tuple[int, date]] = []
        for cell in sheet[2][2:]:
            obs = _quarter_label_to_date(cell.value)
            if obs is None:
                continue
            if start is not None and obs < start:
                continue
            quarter_cols.append((cell.column, obs))

        out: list[CbReserveRow] = []
        for row in sheet.iter_rows(min_row=3):
            country = _cell_text(row[1].value) or _cell_text(row[0].value)
            if not country:
                continue
            iso3 = COUNTRY_ISO3.get(country.lower())
            if iso3 is None:
                logger.debug("wgc_cb: unknown workbook country %r, skipping", country)
                continue
            for col_idx, obs_month in quarter_cols:
                value = row[col_idx - 1].value
                reserves_t = _decimal_or_none(value)
                if reserves_t is None:
                    continue
                out.append(
                    CbReserveRow(
                        country_iso3=iso3,
                        obs_month=obs_month,
                        reserves_t=reserves_t,
                        bucket=classify_bucket(iso3),
                        is_reported=True,
                        is_estimated=False,
                    )
                )
        return out

    def _get_with_telemetry(
        self,
        url: str,
        params: dict[str, Any],
        *,
        endpoint_key: str | None = None,
        endpoint_path: str | None = None,
    ) -> httpx.Response:
        started_at = datetime.now(UTC)
        try:
            response = self._client.get(url, params=params)
        except httpx.HTTPError as exc:
            finished_at = datetime.now(UTC)
            self._record_request(
                self._build_event(
                    started_at,
                    finished_at,
                    params,
                    endpoint_key=endpoint_key,
                    endpoint_path=endpoint_path,
                    status_code=None,
                    error_message=repr(exc)[:1000],
                )
            )
            raise
        finished_at = datetime.now(UTC)
        self._record_request(
            self._build_event(
                started_at,
                finished_at,
                params,
                endpoint_key=endpoint_key,
                endpoint_path=endpoint_path,
                status_code=response.status_code,
                error_message=(
                    response.text[:1000] if response.status_code >= 400 else None
                ),
            )
        )
        return response

    def _record_request(self, event: ExternalApiRequestEvent) -> None:
        if self._record_request_fn is not None:
            self._record_request_fn(self, event)
        else:
            logger.debug("wgc_cb telemetry %r", event)

    def _build_event(
        self,
        started_at: datetime,
        finished_at: datetime,
        params: dict[str, Any],
        *,
        endpoint_key: str | None = None,
        endpoint_path: str | None = None,
        status_code: int | None,
        error_message: str | None,
    ) -> ExternalApiRequestEvent:
        path = endpoint_path or self.ENDPOINT_PATH
        return ExternalApiRequestEvent(
            provider=self.PROVIDER,
            endpoint_key=endpoint_key or self.ENDPOINT_KEY,
            method="GET",
            path=path,
            path_template=path,
            params=redact_params(params),
            status_code=status_code,
            status_family=status_family_for(
                status_code, transport_error=status_code is None
            ),
            started_at=started_at,
            finished_at=finished_at,
            latency_ms=max(0, int((finished_at - started_at).total_seconds() * 1000)),
            error_message=error_message,
        )


def _cell_text(value: object) -> str:
    return str(value).strip() if value is not None else ""


def _decimal_or_none(value: object) -> Decimal | None:
    if value is None or value == "":
        return None
    try:
        return Decimal(str(value))
    except (InvalidOperation, ValueError):
        return None


def _quarter_label_to_date(value: object) -> date | None:
    text = _cell_text(value)
    match = re.fullmatch(r"Q([1-4])\s+(\d{4})", text)
    if match is None:
        return None
    quarter = int(match.group(1))
    year = int(match.group(2))
    month_day = {
        1: (3, 31),
        2: (6, 30),
        3: (9, 30),
        4: (12, 31),
    }[quarter]
    return date(year, *month_day)
