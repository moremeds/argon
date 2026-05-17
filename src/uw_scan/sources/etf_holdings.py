"""Daily ETF holdings for the gold complex.

Targets: GLD (SPDR), IAU (BlackRock), GLDM (SPDR), PHYS (Sprott).
Each fund has its own endpoint and payload shape; we normalise to EtfHoldingRow.
"""

from __future__ import annotations

import csv
import io
import logging
from collections.abc import Callable
from dataclasses import dataclass
from datetime import UTC, date, datetime
from decimal import Decimal, InvalidOperation
from typing import Any

import httpx
from openpyxl import load_workbook

from uw_scan.storage.provider_usage import ExternalApiRequestEvent
from uw_scan.storage.repository import redact_params, status_family_for

logger = logging.getLogger(__name__)


@dataclass(frozen=True)
class EtfHoldingRow:
    ticker: str
    obs_date: date
    holdings_oz: Decimal | None
    shares_out: Decimal | None
    nav_per_share: Decimal | None
    premium_pct: Decimal | None


RecordHook = Callable[["EtfHoldingsProvider", ExternalApiRequestEvent], None]


class EtfHoldingsProvider:
    GLD_URL = "https://api.spdrgoldshares.com/api/v1/historical-archive"
    GLD_PARAMS = {"product": "gld", "exchange": "NYSE", "lang": "en"}
    GLDM_URL = "https://www.spdrgoldshares.com/usa/historical-data-gldm/"
    IAU_URL = "https://www.ishares.com/us/products/239561/iau-holdings.ajax"
    PHYS_URL = "https://sprott.com/api/v1/funds/phys/nav-history"
    PROVIDER = "etf_holdings"

    DEFAULT_TIMEOUT_S = 60.0
    MAX_RETRIES = 3
    BROWSER_UA = (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 13_0) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
    )

    def __init__(
        self,
        *,
        timeout_s: float | None = None,
        max_retries: int | None = None,
        record_request: RecordHook | None = None,
    ):
        self._client = httpx.Client(
            timeout=timeout_s if timeout_s is not None else self.DEFAULT_TIMEOUT_S,
            headers={"User-Agent": self.BROWSER_UA},
        )
        self._max_retries = max_retries if max_retries is not None else self.MAX_RETRIES
        self._record_request_fn = record_request

    def close(self) -> None:
        self._client.close()

    def __enter__(self) -> "EtfHoldingsProvider":
        return self

    def __exit__(self, *_exc: object) -> None:
        self.close()

    def fetch_gld(self, *, start: date | None = None) -> list[EtfHoldingRow]:
        response = self._get_with_telemetry(
            self.GLD_URL, self.GLD_PARAMS, endpoint_key="spdr_gld_archive"
        )
        response.raise_for_status()
        if _looks_like_xlsx(response):
            return self._parse_spdr_archive_xlsx("GLD", response.content, start)
        return self._parse_spdr_csv("GLD", response.text, start)

    def fetch_gldm(self, *, start: date | None = None) -> list[EtfHoldingRow]:
        response = self._get_with_telemetry(
            self.GLDM_URL, {}, endpoint_key="spdr_gldm_csv"
        )
        response.raise_for_status()
        return self._parse_spdr_csv("GLDM", response.text, start)

    def fetch_iau(self, *, start: date | None = None) -> list[EtfHoldingRow]:
        response = self._get_with_telemetry(
            self.IAU_URL, {}, endpoint_key="blackrock_iau"
        )
        response.raise_for_status()
        out: list[EtfHoldingRow] = []
        for row in (response.json() or {}).get("data", []):
            d = _parse_date(row.get("asOfDate"))
            if d is None or (start and d < start):
                continue
            out.append(
                EtfHoldingRow(
                    ticker="IAU",
                    obs_date=d,
                    holdings_oz=_dec(row.get("physicalGoldOunces")),
                    shares_out=None,
                    nav_per_share=_dec(row.get("navPerShare")),
                    premium_pct=None,
                )
            )
        return out

    def fetch_phys(self, *, start: date | None = None) -> list[EtfHoldingRow]:
        response = self._get_with_telemetry(
            self.PHYS_URL, {}, endpoint_key="sprott_phys"
        )
        response.raise_for_status()
        out: list[EtfHoldingRow] = []
        for row in (response.json() or {}).get("data", []):
            d = _parse_date(row.get("date"))
            if d is None or (start and d < start):
                continue
            out.append(
                EtfHoldingRow(
                    ticker="PHYS",
                    obs_date=d,
                    holdings_oz=_dec(row.get("goldOunces")),
                    shares_out=None,
                    nav_per_share=_dec(row.get("nav")),
                    premium_pct=_dec(row.get("premiumDiscountPct")),
                )
            )
        return out

    def _parse_spdr_csv(
        self, ticker: str, text: str, start: date | None
    ) -> list[EtfHoldingRow]:
        out: list[EtfHoldingRow] = []
        reader = csv.DictReader(io.StringIO(text))
        for row in reader:
            d = _parse_date(row.get("Date"))
            if d is None or (start and d < start):
                continue
            out.append(
                EtfHoldingRow(
                    ticker=ticker,
                    obs_date=d,
                    holdings_oz=_dec(row.get("Ounces in the Trust")),
                    shares_out=None,
                    nav_per_share=_dec(row.get("NAV per Share (USD)")),
                    premium_pct=None,
                )
            )
        return out

    def _parse_spdr_archive_xlsx(
        self, ticker: str, content: bytes, start: date | None
    ) -> list[EtfHoldingRow]:
        out: list[EtfHoldingRow] = []
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        sheet_name = f"US {ticker} Historical Archive"
        sheet = workbook[sheet_name] if sheet_name in workbook.sheetnames else None
        if sheet is None:
            logger.warning("spdr archive missing sheet %s", sheet_name)
            return out

        rows = sheet.iter_rows(values_only=True)
        header = next(rows, None)
        if header is None:
            return out
        columns = {str(value).strip(): idx for idx, value in enumerate(header) if value}
        for row in rows:
            d = _parse_date(_cell(row, columns, "Date"))
            if d is None or (start and d < start):
                continue
            holdings_oz = _dec(_cell(row, columns, "Total Ounces of Gold in the Trust"))
            if holdings_oz is None:
                continue
            out.append(
                EtfHoldingRow(
                    ticker=ticker,
                    obs_date=d,
                    holdings_oz=holdings_oz,
                    shares_out=None,
                    nav_per_share=_dec(_cell(row, columns, "NAV/Share at 10:30am NYT")),
                    premium_pct=_dec(
                        _cell(
                            row,
                            columns,
                            "Premium/Discount of GLD Mid Point vs Indicative Value of GLD at 4:15pm NYT",
                        )
                    ),
                )
            )
        return out

    def _get_with_telemetry(
        self, url: str, params: dict[str, Any], *, endpoint_key: str
    ) -> httpx.Response:
        import time

        last_exc: Exception | None = None
        for attempt in range(self._max_retries):
            started_at = datetime.now(UTC)
            try:
                response = self._client.get(url, params=params)
            except (httpx.ReadTimeout, httpx.ConnectTimeout, httpx.ReadError) as exc:
                finished_at = datetime.now(UTC)
                self._record_request(
                    self._build_event(
                        url,
                        params,
                        endpoint_key,
                        started_at,
                        finished_at,
                        status_code=None,
                        error_message=f"attempt {attempt + 1}: {repr(exc)[:900]}",
                    )
                )
                last_exc = exc
                if attempt < self._max_retries - 1:
                    time.sleep(2**attempt)
                    continue
                raise
            except httpx.HTTPError as exc:
                finished_at = datetime.now(UTC)
                self._record_request(
                    self._build_event(
                        url,
                        params,
                        endpoint_key,
                        started_at,
                        finished_at,
                        status_code=None,
                        error_message=repr(exc)[:1000],
                    )
                )
                raise
            finished_at = datetime.now(UTC)
            self._record_request(
                self._build_event(
                    url,
                    params,
                    endpoint_key,
                    started_at,
                    finished_at,
                    status_code=response.status_code,
                    error_message=(
                        response.text[:1000] if response.status_code >= 400 else None
                    ),
                )
            )
            return response
        assert last_exc is not None
        raise last_exc

    def _record_request(self, event: ExternalApiRequestEvent) -> None:
        if self._record_request_fn is not None:
            self._record_request_fn(self, event)
        else:
            logger.debug("etf_holdings telemetry %r", event)

    def _build_event(
        self,
        url: str,
        params: dict[str, Any],
        endpoint_key: str,
        started_at: datetime,
        finished_at: datetime,
        *,
        status_code: int | None,
        error_message: str | None,
    ) -> ExternalApiRequestEvent:
        return ExternalApiRequestEvent(
            provider=self.PROVIDER,
            endpoint_key=endpoint_key,
            method="GET",
            path=url,
            path_template=url,
            params=redact_params(params),
            status_code=status_code,
            status_family=status_family_for(
                status_code, transport_error=status_code is None
            ),
            started_at=started_at,
            finished_at=finished_at,
            latency_ms=max(0, int((finished_at - started_at).total_seconds() * 1000)),
            error_message=error_message,
        )


def _parse_date(raw: str | None) -> date | None:
    if not raw:
        return None
    if isinstance(raw, datetime):
        return raw.date()
    if isinstance(raw, date):
        return raw
    raw = str(raw).strip()
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%d-%b-%Y"):
        try:
            return datetime.strptime(raw, fmt).date()
        except ValueError as exc:
            logger.debug("etf date parse fmt=%s skipped: %s", fmt, repr(exc))
            continue
    return None


def _looks_like_xlsx(response: httpx.Response) -> bool:
    content_type = response.headers.get("content-type", "").lower()
    return (
        response.content.startswith(b"PK\x03\x04")
        or "spreadsheetml.sheet" in content_type
    )


def _cell(row: tuple[Any, ...], columns: dict[str, int], name: str) -> Any:
    idx = columns.get(name)
    if idx is None or idx >= len(row):
        return None
    return row[idx]


def _dec(raw: Any) -> Decimal | None:
    if raw is None or raw == "":
        return None
    try:
        return Decimal(str(raw).replace(",", ""))
    except (InvalidOperation, ValueError) as exc:
        logger.debug("etf decimal parse skipped: %s", repr(exc))
        return None
