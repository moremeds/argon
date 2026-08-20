"""Official New York Fed Survey of Market Expectations policy source.

The publisher supplies both a structured XLSX and a human-readable PDF.  The
XLSX is the critical data path because it preserves panel identity, publisher
value tags, horizons, units, and aggregation types without PDF-coordinate
inference.  The exact PDF remains attached to the source bundle for review.
"""

from __future__ import annotations

import logging
import re
from collections import defaultdict
from collections.abc import Collection
from dataclasses import dataclass
from datetime import UTC, date, datetime, timedelta
from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Final, Literal
from urllib.parse import urljoin

import httpx
from bs4 import BeautifulSoup
from openpyxl import load_workbook

from uw_scan.macro_evidence import macro_artifact_content_identity
from uw_scan.models.macro import MacroSourceArtifact
from uw_scan.normalize import NormalizationError

logger = logging.getLogger(__name__)

PARSER_VERSION: Final = "nyfed_sme.v1"
SOURCE: Final = "new_york_fed_sme"
SmePanelType = Literal["Dealer", "Participant", "Combined"]

_REQUIRED_COLUMNS: Final = {
    "survey_release_date",
    "survey_due_date",
    "panel_type",
    "spd_question_number",
    "theme",
    "subject_group",
    "subject",
    "question_type",
    "question_mode",
    "question_text",
    "question_tag",
    "value_tag",
    "top_header_value",
    "left_header_value",
    "horizon",
    "horizon_date",
    "bucket_range",
    "bucket_low",
    "bucket_high",
    "aggregation",
    "aggregation_value",
}


@dataclass(frozen=True)
class SmePathPoint:
    horizon: str
    horizon_date: date | None
    respondent_count: int
    p25: Decimal
    median: Decimal
    p75: Decimal
    unit: str
    source_record_id: str


@dataclass(frozen=True)
class SmeProbabilityBucket:
    label: str
    lower_bound: Decimal | None
    upper_bound: Decimal | None
    probability: Decimal
    unit: str = "percent"


@dataclass(frozen=True)
class SmeProbabilityDistribution:
    horizon: str
    horizon_date: date | None
    respondent_count: int
    buckets: tuple[SmeProbabilityBucket, ...]
    source_record_id: str


@dataclass(frozen=True)
class SmeRelease:
    survey_release_date: date
    response_due_date: date
    published_at: datetime | None
    available_at: datetime
    panel_type: SmePanelType
    source_url: str
    source_record_id: str
    path_points: tuple[SmePathPoint, ...]
    probability_distributions: tuple[SmeProbabilityDistribution, ...]


@dataclass(frozen=True)
class SmeLinks:
    survey_month: date
    data_url: str
    report_url: str


@dataclass(frozen=True)
class SmeSourceBundle:
    survey_month: date
    data_artifact: MacroSourceArtifact
    report_artifact: MacroSourceArtifact

    @classmethod
    def from_bytes(
        cls,
        *,
        survey_month: date,
        data_url: str,
        data_bytes: bytes,
        report_url: str,
        report_bytes: bytes,
        retrieved_at: datetime,
    ) -> "SmeSourceBundle":
        month_start = survey_month.replace(day=1)
        record_base = f"nyfed-sme:{month_start:%Y-%m}"
        return cls(
            survey_month=month_start,
            data_artifact=_artifact(
                source_record_id=f"{record_base}:xlsx",
                source_url=data_url,
                media_type=(
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                ),
                raw_bytes=data_bytes,
                retrieved_at=retrieved_at,
            ),
            report_artifact=_artifact(
                source_record_id=f"{record_base}:pdf",
                source_url=report_url,
                media_type="application/pdf",
                raw_bytes=report_bytes,
                retrieved_at=retrieved_at,
            ),
        )


class NyFedSmeProvider:
    BASE_URL = "https://www.newyorkfed.org"
    LANDING_PATH = "/markets/market-intelligence/survey-of-market-expectations"

    def __init__(
        self,
        *,
        base_url: str = BASE_URL,
        timeout_s: float = 30.0,
        trust_env: bool = False,
    ):
        self._base_url = base_url.rstrip("/")
        self._client = httpx.Client(
            timeout=timeout_s,
            follow_redirects=True,
            trust_env=trust_env,
        )

    def __enter__(self) -> "NyFedSmeProvider":
        return self

    def __exit__(self, *_exc: object) -> None:
        self.close()

    def close(self) -> None:
        self._client.close()

    def fetch_latest_bundle(
        self, *, retrieved_at: datetime | None = None
    ) -> SmeSourceBundle:
        return self.fetch_bundles(retrieved_at=retrieved_at)[-1]

    def list_survey_months(self) -> tuple[date, ...]:
        """Every survey month the publisher currently lists, oldest first.

        Cheap next to a bundle fetch (one HTML GET, no XLSX/PDF), so a caller can
        ask what exists before deciding what to download.
        """
        return tuple(links.survey_month for links in self._discover())

    def _discover(self) -> tuple[SmeLinks, ...]:
        landing = self._get(self.LANDING_PATH)
        landing.raise_for_status()
        return discover_sme_links(landing.content, base_url=self._base_url)

    def fetch_bundles(
        self,
        *,
        survey_months: Collection[date] | None = None,
        retrieved_at: datetime | None = None,
    ) -> tuple[SmeSourceBundle, ...]:
        """Download one bundle per requested survey month, oldest first.

        The landing page has always listed every survey the NY Fed still hosts --
        ``discover_sme_links`` returns all of them and the nightly job simply took
        ``[-1]``.  So the history was never unreachable; it was unasked for.

        ``survey_months=None`` keeps that nightly contract exactly: latest only.
        A requested month the publisher does not list is an error naming the month
        and what is on offer, never a silent short read -- a backfill that quietly
        returns 11 of 12 surveys reports success for a hole it just made.
        """
        available = self._discover()
        if survey_months is None:
            wanted = available[-1:]
        else:
            starts = {month.replace(day=1) for month in survey_months}
            by_month = {links.survey_month: links for links in available}
            missing = sorted(starts - by_month.keys())
            if missing:
                raise NormalizationError(
                    "NY Fed SME publisher page does not list survey month(s) "
                    f"{', '.join(f'{month:%Y-%m}' for month in missing)}; "
                    f"available: {', '.join(f'{m:%Y-%m}' for m in by_month)}"
                )
            wanted = tuple(by_month[month] for month in sorted(starts))
        seen_at = retrieved_at or datetime.now(UTC)
        return tuple(self._bundle_for(links, retrieved_at=seen_at) for links in wanted)

    def _bundle_for(
        self, links: SmeLinks, *, retrieved_at: datetime
    ) -> SmeSourceBundle:
        data_response = self._get(links.data_url)
        data_response.raise_for_status()
        report_response = self._get(links.report_url)
        report_response.raise_for_status()
        return SmeSourceBundle.from_bytes(
            survey_month=links.survey_month,
            data_url=links.data_url,
            data_bytes=data_response.content,
            report_url=links.report_url,
            report_bytes=report_response.content,
            retrieved_at=retrieved_at,
        )

    def _get(self, path_or_url: str) -> httpx.Response:
        url = (
            path_or_url
            if path_or_url.startswith("http")
            else f"{self._base_url}{path_or_url}"
        )
        return self._client.get(url)


def discover_sme_links(raw_html: bytes, *, base_url: str) -> tuple[SmeLinks, ...]:
    soup = BeautifulSoup(raw_html, "html.parser")
    data_urls: dict[date, str] = {}
    report_urls: dict[date, str] = {}
    months = {
        "jan": 1,
        "feb": 2,
        "mar": 3,
        "apr": 4,
        "may": 5,
        "jun": 6,
        "jul": 7,
        "aug": 8,
        "sep": 9,
        "oct": 10,
        "nov": 11,
        "dec": 12,
    }
    for anchor in soup.find_all("a", href=True):
        href = str(anchor["href"])
        data_match = re.search(r"/([a-z]{3})-(\d{4})-data\.xlsx$", href)
        report_match = re.search(r"/([a-z]{3})-(\d{4})-sme-results\.pdf$", href)
        match = data_match or report_match
        if match is None or match.group(1) not in months:
            continue
        survey_month = date(int(match.group(2)), months[match.group(1)], 1)
        absolute = urljoin(f"{base_url.rstrip('/')}/", href)
        if data_match is not None:
            data_urls[survey_month] = absolute
        else:
            report_urls[survey_month] = absolute
    paired = sorted(data_urls.keys() & report_urls.keys())
    if not paired:
        raise NormalizationError(
            "NY Fed SME publisher page has no paired XLSX/PDF releases"
        )
    return tuple(
        SmeLinks(
            survey_month=survey_month,
            data_url=data_urls[survey_month],
            report_url=report_urls[survey_month],
        )
        for survey_month in paired
    )


def parse_sme_release(
    bundle: SmeSourceBundle, *, panel_type: SmePanelType = "Dealer"
) -> SmeRelease:
    raw = bundle.data_artifact.raw_bytes
    if raw is None:
        raise NormalizationError("NY Fed SME data artifact is missing raw bytes")
    rows, sheet_name = _workbook_rows(raw)
    filtered = [
        row
        for row in rows
        if row["panel_type"] == panel_type
        and row["subject_group"] == "fed_funds_target_range"
    ]
    if not filtered:
        raise NormalizationError(f"NY Fed SME workbook has no {panel_type} policy rows")
    if any(row["question_mode"] != "levels" for row in filtered):
        raise NormalizationError("NY Fed SME policy units/mode changed from levels")

    release_dates = {_date(row["survey_release_date"]) for row in filtered}
    due_dates = {_date(row["survey_due_date"]) for row in filtered}
    if len(release_dates) != 1 or len(due_dates) != 1:
        raise NormalizationError("NY Fed SME survey dates are inconsistent")
    release_date = next(iter(release_dates))
    due_date = next(iter(due_dates))
    # The filename month labels the FOMC cycle the survey feeds; the workbook's own
    # release date is when it was actually published, and the two are not required to
    # agree.  Measured across the 12 hosted surveys, 10 match and 2 publish in the
    # PRIOR month (may-2025 released 2025-04-23, dec-2025 released 2025-11-25), so
    # demanding equality rejected real releases for following the publisher's calendar.
    #
    # Still fails closed on the mispairing this guards against: fetching one month's
    # workbook under another's key is months out, not weeks, and is caught here.
    labelled = bundle.survey_month.replace(day=1)
    earliest = (labelled - timedelta(days=1)).replace(day=1)
    if not earliest <= release_date.replace(day=1) <= labelled:
        raise NormalizationError(
            f"NY Fed SME workbook release date {release_date} is outside the "
            f"{labelled:%Y-%m} release window (that month or the one before it)"
        )

    path_points = _path_points(filtered, sheet_name=sheet_name, panel_type=panel_type)
    distributions = _probability_distributions(
        filtered,
        sheet_name=sheet_name,
        panel_type=panel_type,
    )
    if not path_points:
        raise NormalizationError("NY Fed SME dealer policy path is empty")
    if not distributions:
        raise NormalizationError(
            "NY Fed SME dealer probability distributions are empty"
        )

    record_base = f"nyfed-sme:{bundle.survey_month:%Y-%m}"
    return SmeRelease(
        survey_release_date=release_date,
        response_due_date=due_date,
        # Deliberately None, not midnight on ``release_date``.  The NY Fed states a
        # release DATE; ``published_at`` is an aware instant, so deriving one would
        # invent an hour the publisher never gave and let a replay claim the survey
        # was readable before it was.  ``response_due_date`` carries the real survey
        # date on the observation, and that is what orders releases.
        published_at=None,
        available_at=bundle.data_artifact.available_at,
        panel_type=panel_type,
        source_url=bundle.data_artifact.source_url or "",
        source_record_id=f"{record_base}:{panel_type}",
        path_points=path_points,
        probability_distributions=distributions,
    )


def _artifact(
    *,
    source_record_id: str,
    source_url: str,
    media_type: str,
    raw_bytes: bytes,
    retrieved_at: datetime,
) -> MacroSourceArtifact:
    content_hash, content_length = macro_artifact_content_identity(raw_bytes=raw_bytes)
    return MacroSourceArtifact(
        source=SOURCE,
        source_kind="official",
        source_record_id=source_record_id,
        source_url=source_url,
        published_at=None,
        available_at=retrieved_at,
        retrieved_at=retrieved_at,
        last_seen_at=retrieved_at,
        content_hash=content_hash,
        parser_version=PARSER_VERSION,
        quality_status="partial",
        cost_class="free_official",
        media_type=media_type,
        content_length=content_length,
        raw_bytes=raw_bytes,
    )


def _workbook_rows(raw: bytes) -> tuple[list[dict[str, object]], str]:
    try:
        workbook = load_workbook(BytesIO(raw), read_only=True, data_only=True)
    except Exception as exc:
        raise NormalizationError(
            f"NY Fed SME workbook cannot be opened: {exc!r}"
        ) from exc
    try:
        sheet = workbook.active
        iterator = sheet.iter_rows(values_only=True)
        header_row = next(iterator, None)
        if header_row is None:
            raise NormalizationError("NY Fed SME workbook is empty")
        headers = tuple(str(value) if value is not None else "" for value in header_row)
        missing = sorted(_REQUIRED_COLUMNS - set(headers))
        if missing:
            raise NormalizationError(
                f"NY Fed SME workbook missing required columns: {missing}"
            )
        rows = [dict(zip(headers, row, strict=True)) for row in iterator]
        return rows, sheet.title
    finally:
        workbook.close()


def _path_points(
    rows: list[dict[str, object]],
    *,
    sheet_name: str,
    panel_type: SmePanelType,
) -> tuple[SmePathPoint, ...]:
    selected = [row for row in rows if row["question_type"] == "path_of_modes"]
    if selected and {str(row["question_tag"]) for row in selected} != {
        "fftr_pathofmodes"
    }:
        raise NormalizationError("NY Fed SME policy path label changed")
    grouped: dict[str, list[dict[str, object]]] = defaultdict(list)
    for row in selected:
        grouped[str(row["value_tag"])].append(row)

    points: list[SmePathPoint] = []
    for value_tag, group in grouped.items():
        aggregations = _aggregation_map(group)
        required = {"count", "pctl25", "pctl50", "pctl75"}
        if set(aggregations) != required:
            raise NormalizationError(
                f"NY Fed SME {value_tag} aggregations changed: {sorted(aggregations)}"
            )
        first = group[0]
        points.append(
            SmePathPoint(
                horizon=str(first["horizon"]),
                horizon_date=_optional_date(first["horizon_date"]),
                respondent_count=_count(aggregations["count"], context=value_tag),
                p25=_percent(aggregations["pctl25"], context=value_tag),
                median=_percent(aggregations["pctl50"], context=value_tag),
                p75=_percent(aggregations["pctl75"], context=value_tag),
                unit="percent",
                source_record_id=(
                    f"nyfed-sme:{_date(first['survey_release_date']):%Y-%m}:"
                    f"{sheet_name}:{panel_type}:{value_tag}"
                ),
            )
        )
    return tuple(
        sorted(
            points, key=lambda point: (point.horizon_date or date.max, point.horizon)
        )
    )


def _probability_distributions(
    rows: list[dict[str, object]],
    *,
    sheet_name: str,
    panel_type: SmePanelType,
) -> tuple[SmeProbabilityDistribution, ...]:
    selected = [
        row for row in rows if row["question_type"] == "probability_distribution"
    ]
    grouped: dict[str, list[dict[str, object]]] = defaultdict(list)
    for row in selected:
        grouped[str(row["question_tag"])].append(row)

    output: list[SmeProbabilityDistribution] = []
    for question_tag, group in grouped.items():
        bucket_groups: dict[str, list[dict[str, object]]] = defaultdict(list)
        for row in group:
            bucket_groups[str(row["value_tag"])].append(row)
        buckets: list[SmeProbabilityBucket] = []
        counts: set[int] = set()
        first = group[0]
        for value_tag, bucket_group in bucket_groups.items():
            aggregations = _aggregation_map(bucket_group)
            if set(aggregations) != {"count", "avg"}:
                raise NormalizationError(
                    f"NY Fed SME {value_tag} probability aggregations changed"
                )
            counts.add(_count(aggregations["count"], context=value_tag))
            bucket_row = bucket_group[0]
            buckets.append(
                SmeProbabilityBucket(
                    label=str(bucket_row["bucket_range"]),
                    lower_bound=_optional_percent(
                        bucket_row["bucket_low"], context=value_tag
                    ),
                    upper_bound=_optional_percent(
                        bucket_row["bucket_high"], context=value_tag
                    ),
                    probability=_percent(aggregations["avg"], context=value_tag),
                )
            )
        if len(counts) != 1:
            raise NormalizationError(
                f"NY Fed SME {question_tag} respondent counts are inconsistent"
            )
        total = sum((bucket.probability for bucket in buckets), Decimal(0))
        # Each bucket is a mean across respondents that the NY Fed publishes already
        # rounded, so a correct distribution does NOT sum to exactly 100 -- N buckets
        # rounded to the whole percent can be off by up to 0.5 each.  The tolerance is
        # therefore derived from the bucket count rather than fixed: it is the tightest
        # bound that cannot reject a correctly-rounded release.
        #
        # A fixed +/-1 band was rejecting real data.  Measured over all 12 surveys the
        # publisher currently hosts (2025-01..2026-06, ~13 distributions each), totals
        # run 98..102 -- so the band threw away 6 of 12 surveys, and the failure named
        # the probability sub-table while what it actually cost was the policy path.
        # A dropped bucket, which is the parse error this guard exists to catch, moves
        # a 10-bucket total by ~10 and is still caught with room to spare.
        tolerance = Decimal("0.5") * len(buckets)
        if abs(total - Decimal(100)) > tolerance:
            raise NormalizationError(
                f"NY Fed SME {question_tag} probability total {total} is further than "
                f"{tolerance} from 100 across {len(buckets)} rounded buckets"
            )
        output.append(
            SmeProbabilityDistribution(
                horizon=str(first["horizon"]),
                horizon_date=_optional_date(first["horizon_date"]),
                respondent_count=next(iter(counts)),
                buckets=tuple(buckets),
                source_record_id=(
                    f"nyfed-sme:{_date(first['survey_release_date']):%Y-%m}:"
                    f"{sheet_name}:{panel_type}:{question_tag}"
                ),
            )
        )
    return tuple(
        sorted(output, key=lambda item: (item.horizon_date or date.max, item.horizon))
    )


def _aggregation_map(group: list[dict[str, object]]) -> dict[str, object]:
    output: dict[str, object] = {}
    for row in group:
        key = str(row["aggregation"])
        if key in output:
            raise NormalizationError(f"NY Fed SME duplicate {key} aggregation")
        output[key] = row["aggregation_value"]
    return output


def _date(raw: object) -> date:
    if isinstance(raw, datetime):
        return raw.date()
    if isinstance(raw, date):
        return raw
    try:
        return date.fromisoformat(str(raw))
    except ValueError as exc:
        raise NormalizationError(f"NY Fed SME date is invalid: {raw!r}") from exc


def _optional_date(raw: object) -> date | None:
    return None if raw in (None, "") else _date(raw)


def _decimal(raw: object, *, context: str) -> Decimal:
    try:
        value = Decimal(str(raw))
    except (InvalidOperation, ValueError) as exc:
        raise NormalizationError(f"NY Fed SME {context} is not numeric") from exc
    if not value.is_finite():
        raise NormalizationError(f"NY Fed SME {context} must be finite")
    return value


def _percent(raw: object, *, context: str) -> Decimal:
    return _decimal(raw, context=context) * Decimal(100)


def _optional_percent(raw: object, *, context: str) -> Decimal | None:
    return None if raw in (None, "") else _percent(raw, context=context)


def _count(raw: object, *, context: str) -> int:
    value = _decimal(raw, context=context)
    if value != value.to_integral_value() or value < 0:
        raise NormalizationError(
            f"NY Fed SME {context} count is not nonnegative integer"
        )
    return int(value)
